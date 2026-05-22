#!/usr/bin/env python3
"""
coletor_pedal.py — descompacta exports do WhatsApp e arquiva as mídias nas
pastas corretas dos passeios do Pedal Hidrográfico, renomeando com o handle
do Instagram de quem enviou.

Como gerar o zip de entrada:
  1. WhatsApp → grupo do Pedal → ⋮ → "Exportar conversa" → INCLUIR MÍDIA
  2. Salvar/transferir o .zip para a pasta de exports (default ~/Downloads/PedalHidrografico-exports/)
  3. Rodar este script.

Uso:
  python3 coletor_pedal.py                       # processa todos zips na pasta default
  python3 coletor_pedal.py arq1.zip arq2.zip     # processa zips específicos
  python3 coletor_pedal.py --watch               # roda continuamente, polling cada 15s
  python3 coletor_pedal.py --pedais /caminho     # pasta-mãe dos passeios
  python3 coletor_pedal.py --exports /caminho    # pasta dos zips
  python3 coletor_pedal.py --dry-run             # mostra o que faria, sem mover nada
  python3 coletor_pedal.py --max-days 1          # janela de "dia seguinte"

Dependências: só stdlib. Python 3.9+.
"""

import argparse
import json
import re
import shutil
import sys
import time
import zipfile
from datetime import datetime, timedelta
from pathlib import Path

# ─── Defaults (sobreescritos pelos flags da linha de comando) ───────────────
EXPORTS_DIR_DEFAULT = Path.home() / "Downloads" / "PedalHidrografico-exports"
PEDAL_DIR_DEFAULT   = Path("/Users/danlessa/pedalhidro-teste")
PROCESSED_SUBDIR    = "processed"
HANDLES_FILE        = Path(__file__).resolve().parent / "coletor_pedal_handles.json"
# Planilha-fonte (xlsx) — quando informada via --xlsx, vira a fonte de verdade
# sobre quando cada pedal aconteceu e qual seu eH/nome.
XLSX_DEFAULT        = Path.home() / "Downloads" / "Censo Hidrográfico - microdados.xlsx"
# Quantos dias DEPOIS do pedal as fotos ainda contam pra ele. Pedais são
# semanais, então 1-2 é seguro; acima disso o script cria uma pasta nova.
MAX_DAYS_AFTER_PEDAL_DEFAULT = 1

# ─── Parser do _chat.txt do WhatsApp ────────────────────────────────────────
# Formatos suportados (WhatsApp varia entre versões iOS/Android/anos):
#   [15/05/26, 00:04:38] Fulano: <attached: arquivo.jpg>     ← iOS recente
#   [09/05/2026, 19:42:13] Fulano: arquivo.jpg (file attached) ← iOS antigo
#   09/05/2026 19:42 - Fulano: arquivo.jpg (arquivo anexado)   ← Android
HEADER_IOS = re.compile(
    r'^\[(\d{2}/\d{2}/\d{2,4}),?\s*(\d{1,2}:\d{2}(?::\d{2})?)\]\s*([^:]+):\s*(.*)'
)
HEADER_ANDROID = re.compile(
    r'^(\d{2}/\d{2}/\d{2,4})\s+(\d{1,2}:\d{2}(?::\d{2})?)\s+-\s+([^:]+):\s*(.*)'
)
# Marcadores de anexo:
#   "<attached: arquivo.jpg>"                         ← iOS recente
#   "arquivo.jpg (arquivo anexado)" / "(file attached)" ← antigo
ATTACHED_ANGLE = re.compile(r'<attached:\s*([^>]+)>', re.IGNORECASE)
ATTACHED_PAREN = re.compile(
    r'^(.+?)\s+\((?:arquivo anexado|file attached)\)\s*$', re.IGNORECASE
)
# Caracteres invisíveis que o WhatsApp gosta de meter no começo das mensagens.
INVISIBLE = '‎‏‪‬﻿'

def _parse_dt(date_s, time_s):
    """dd/mm/yy(yy) + hh:mm[:ss] → datetime. Aceita 2 ou 4 dígitos no ano."""
    year_fmt = "%Y" if len(date_s.split('/')[2]) == 4 else "%y"
    time_fmt = "%H:%M:%S" if time_s.count(':') == 2 else "%H:%M"
    return datetime.strptime(f"{date_s} {time_s}", f"%d/%m/{year_fmt} {time_fmt}")

def parse_chat(chat_text):
    """Devolve lista [(datetime, sender, attached_filename_or_None, raw_msg)]."""
    entries = []
    current = None
    for line in chat_text.splitlines():
        for ch in INVISIBLE:
            line = line.replace(ch, '')
        m = HEADER_IOS.match(line) or HEADER_ANDROID.match(line)
        if m:
            date_s, time_s, sender, msg = m.groups()
            try:
                dt = _parse_dt(date_s, time_s)
            except ValueError:
                continue
            current = [dt, sender.strip(), None, msg]
            entries.append(current)
        elif current is not None:
            current[3] += "\n" + line

    for entry in entries:
        text = entry[3].strip()
        # iOS recente: "<attached: filename>"
        m = ATTACHED_ANGLE.search(text)
        if m:
            entry[2] = m.group(1).strip()
            continue
        # Antigo: "filename (arquivo anexado)" / "(file attached)"
        m = ATTACHED_PAREN.match(text)
        if m:
            entry[2] = m.group(1).strip()
    return entries

# ─── Descoberta das pastas de passeios ──────────────────────────────────────
FOLDER_NAME_RE = re.compile(r'^(\d{4}-\d{2}-\d{2})\s*(\d+)?\s*(.*)$')

def discover_pedal_folders(root):
    """Lista [(date, numero_int_ou_None, Path)] ordenada por data."""
    out = []
    if not root.exists():
        return out
    for p in root.iterdir():
        if not p.is_dir() or p.name.startswith('.'):
            continue
        m = FOLDER_NAME_RE.match(p.name)
        if not m:
            continue
        try:
            d = datetime.strptime(m.group(1), "%Y-%m-%d").date()
        except ValueError:
            continue
        n = int(m.group(2)) if m.group(2) and m.group(2).isdigit() else None
        out.append((d, n, p))
    out.sort(key=lambda t: t[0])
    return out

def next_pedal_number(existing, sheet_pedais=None):
    """Maior número conhecido + 1, considerando pastas existentes E planilha."""
    nums = [n for _, n, _ in existing if n is not None]
    if sheet_pedais:
        nums += [eh for _, eh, _ in sheet_pedais if eh is not None]
    return max(nums) + 1 if nums else 1

def sanitize_folder_name(s):
    """Remove caracteres problemáticos pra nome de pasta no macOS/Linux."""
    s = (s or '').strip()
    # / e : são especiais; deixar mais portátil
    return s.replace('/', '-').replace(':', '-').strip() or '(sem nome)'

def load_pedais_from_xlsx(xlsx_path):
    """Lê a aba 'Geral' e retorna [(date, eH, nome), ...] ordenado por data.
    Retorna [] em caso de erro (e printa aviso). openpyxl é opcional —
    só precisa estar instalado se --xlsx for usado.
    """
    try:
        import openpyxl
    except ImportError:
        print(f"⚠️  openpyxl não instalado — `pip3 install openpyxl` pra usar --xlsx",
              file=sys.stderr)
        return []
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        print(f"⚠️  não consegui abrir {xlsx_path}: {e}", file=sys.stderr)
        return []
    if "Geral" not in wb.sheetnames:
        print(f"⚠️  {xlsx_path} sem aba 'Geral'", file=sys.stderr)
        return []
    ws = wb["Geral"]
    # Cols (1-based): A=Data, B=eH, G=Nome
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row: continue
        data, eh, nome = row[0], row[1], row[6] if len(row) > 6 else None
        if not data or eh is None: continue
        # Data pode vir como datetime, date ou string
        if hasattr(data, 'date'):
            d = data.date()
        elif hasattr(data, 'year'):
            d = data
        elif isinstance(data, str):
            try:
                d = datetime.strptime(data[:10], "%Y-%m-%d").date()
            except ValueError:
                continue
        else:
            continue
        try:
            eh_int = int(eh)
        except (ValueError, TypeError):
            continue
        out.append((d, eh_int, nome or ''))
    out.sort(key=lambda t: t[0])
    return out

def find_pedal_in_sheet(photo_date, sheet_pedais, max_days):
    """Acha na planilha o pedal mais recente cuja data <= photo_date dentro
    da janela max_days. Retorna (date, eH, nome) ou None."""
    if not sheet_pedais:
        return None
    for d, eh, nome in reversed(sheet_pedais):
        if d <= photo_date:
            age = (photo_date - d).days
            return (d, eh, nome) if age <= max_days else None
    return None

def find_or_create_pedal_folder(root, photo_dt, existing, sheet_pedais,
                                max_days, dry_run):
    """
    Só processa fotos que casam com algum pedal da planilha (data ≤ photo_date,
    idade ≤ max_days). Sem match → retorna (None, motivo) e o caller pula.

    Estratégia:
      1. Se planilha tem pedal cuja data <= photo_date e idade <= max_days,
         usa esse (eH + nome corretos) — procura/cria pasta correspondente.
      2. Senão, NÃO cria pasta e NÃO processa. Retorna (None, motivo).
    """
    photo_date = photo_dt.date()

    sheet_match = find_pedal_in_sheet(photo_date, sheet_pedais, max_days)
    if sheet_match is None:
        return None, "sem pedal na planilha pra essa data"

    s_date, s_eh, s_nome = sheet_match
    # Existe pasta com mesma data OU mesmo número?
    for d, n, p in existing:
        if d == s_date or (n is not None and n == s_eh):
            if not dry_run:
                (p / "midia").mkdir(exist_ok=True)
            return p, f"planilha eH={s_eh} (pasta existente)"
    # Cria pasta nova com nome canônico
    folder_name = f"{s_date.isoformat()} {s_eh} {sanitize_folder_name(s_nome)}"
    new = root / folder_name
    if not dry_run:
        (new / "midia").mkdir(parents=True, exist_ok=True)
    return new, f"NOVO da planilha eH={s_eh}"

# ─── Handles ────────────────────────────────────────────────────────────────
def load_handles(extra_path=None):
    handles = {}
    if HANDLES_FILE.exists():
        try:
            handles.update(json.loads(HANDLES_FILE.read_text(encoding='utf-8')))
        except Exception as e:
            print(f"⚠️  {HANDLES_FILE} inválido: {e}", file=sys.stderr)
    if extra_path and Path(extra_path).exists():
        try:
            handles.update(json.loads(Path(extra_path).read_text(encoding='utf-8')))
        except Exception as e:
            print(f"⚠️  {extra_path} inválido: {e}", file=sys.stderr)
    return handles

def resolve_handle(sender_name, handles, unknown_log):
    """Encontra o handle IG para um nome do WhatsApp.
    Tenta match exato; depois match case-insensitive. Loga quem não encontrou.
    """
    if sender_name in handles:
        return handles[sender_name]
    lower = {k.lower(): v for k, v in handles.items()}
    if sender_name.lower() in lower:
        return lower[sender_name.lower()]
    unknown_log.add(sender_name)
    return ''  # vazio = sem prefixo, mantém o nome original do arquivo

# ─── Processamento principal ────────────────────────────────────────────────
def process_zip(zip_path, root, exports_dir, handles, sheet_pedais,
                max_days, dry_run, archive_zip):
    print(f"\n{'='*72}\n📦 {zip_path.name}\n{'='*72}")
    with zipfile.ZipFile(zip_path, 'r') as z:
        names = z.namelist()
        # _chat.txt pode estar com nomes diferentes (_chat.txt, WhatsApp Chat with X.txt)
        chat_file = next(
            (n for n in names if n.lower().endswith('.txt') and 'chat' in n.lower()),
            None,
        ) or next((n for n in names if n.endswith('.txt')), None)
        if not chat_file:
            print("  ⚠️  Sem .txt de conversa — pulando")
            return

        try:
            chat_text = z.read(chat_file).decode('utf-8')
        except UnicodeDecodeError:
            chat_text = z.read(chat_file).decode('utf-8', errors='replace')
        entries = parse_chat(chat_text)

        # Mapa filename → (datetime, sender). Primeira ocorrência ganha.
        file_meta = {}
        for dt, sender, fn, _ in entries:
            if fn and fn not in file_meta:
                file_meta[fn] = (dt, sender)

        existing = discover_pedal_folders(root)
        moved, skipped, orphan, no_pedal_match = 0, 0, 0, 0
        per_pedal = {}
        unknown_senders = set()
        no_pedal_examples = []  # (date, sender, filename) — primeiros casos

        for name in names:
            base = Path(name).name
            if not base or name.endswith('/'):
                continue
            if name == chat_file or base == chat_file or base.lower().startswith('icon'):
                continue
            if base.startswith('.'):
                continue  # macOS Icon files, etc.

            meta = file_meta.get(base)
            if not meta:
                # Pode ser que o caption tenha um nome ligeiramente diferente
                # (alguns exports renomeiam ao zipar). Skip orphan with notice.
                if base.lower().endswith('.txt'):
                    continue
                print(f"  ⚠️  órfão (sem mensagem): {base}")
                orphan += 1
                continue

            dt, sender = meta
            handle = resolve_handle(sender, handles, unknown_senders)
            folder, status = find_or_create_pedal_folder(
                root, dt, existing, sheet_pedais, max_days, dry_run
            )
            if folder is None:
                # Sem pedal correspondente na planilha — pula sem criar pasta
                no_pedal_match += 1
                if len(no_pedal_examples) < 5:
                    no_pedal_examples.append((dt.date().isoformat(), sender, base))
                continue
            # Refresh listing se acabou de criar
            if folder not in [pp for _, _, pp in existing]:
                existing = discover_pedal_folders(root) if not dry_run else \
                           existing + [(dt.date(), None, folder)]

            dest_name = f"{handle} {base}" if handle else base
            dest = folder / "midia" / dest_name

            if dest.exists():
                skipped += 1
                continue

            if dry_run:
                moved += 1
            else:
                with z.open(name) as src, open(dest, 'wb') as f:
                    shutil.copyfileobj(src, f)
                moved += 1
            per_pedal.setdefault(folder.name, []).append((dest_name, status, sender))

    # Resumo
    verb = "ARQUIVARIA" if dry_run else "arquivado"
    print(f"\n  ✅ {moved} {verb} · {skipped} já existiam · "
          f"{orphan} órfãos · {no_pedal_match} sem pedal na planilha")
    for pedal, items in sorted(per_pedal.items()):
        senders = sorted({s for _, _, s in items})
        print(f"    → {pedal}/midia/  ({len(items)} arquivos, "
              f"de {', '.join(senders)})")
    if no_pedal_match:
        print(f"\n  ⏭️  {no_pedal_match} arquivos pulados (data não bate com "
              f"nenhum pedal da planilha, dentro de {max_days}d):")
        for d, s, fn in no_pedal_examples:
            print(f"      - {d}  {s[:30]:<30}  {fn}")
        if no_pedal_match > len(no_pedal_examples):
            print(f"      … e mais {no_pedal_match - len(no_pedal_examples)}")
    if unknown_senders:
        print(f"\n  ⚠️  Remetentes sem handle conhecido (sem prefixo no nome):")
        for s in sorted(unknown_senders):
            print(f"      - {s}")
        print("      Adicione em coletor_pedal_handles.json para corrigir.")

    # Arquiva o zip processado (opt-in). Sem --archive-zip o zip fica onde está,
    # e numa eventual reexecução cada arquivo individual é pulado pelo dedup
    # (já existe no destino), então é seguro reprocessar.
    if archive_zip and not dry_run:
        processed_dir = exports_dir / PROCESSED_SUBDIR
        processed_dir.mkdir(exist_ok=True)
        target = processed_dir / zip_path.name
        if target.exists():
            target = processed_dir / f"{zip_path.stem}-{int(time.time())}.zip"
        zip_path.rename(target)

# ─── CLI ────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("zips", nargs='*', help="Zips específicos para processar")
    ap.add_argument("--watch", action='store_true',
                    help="Roda continuamente, processando zips novos a cada 15s")
    ap.add_argument("--exports", type=Path, default=EXPORTS_DIR_DEFAULT,
                    help=f"Pasta dos zips (default: {EXPORTS_DIR_DEFAULT})")
    ap.add_argument("--pedais", type=Path, default=PEDAL_DIR_DEFAULT,
                    help=f"Pasta-mãe dos passeios (default: {PEDAL_DIR_DEFAULT})")
    ap.add_argument("--max-days", type=int, default=MAX_DAYS_AFTER_PEDAL_DEFAULT,
                    help="Dias depois do pedal em que fotos ainda contam pra ele "
                         f"(default: {MAX_DAYS_AFTER_PEDAL_DEFAULT})")
    ap.add_argument("--handles", type=Path,
                    help="JSON adicional com mapeamento sender→handle")
    ap.add_argument("--xlsx", type=Path, default=XLSX_DEFAULT,
                    help=f"Planilha-fonte (aba Geral) com Data/eH/Nome dos "
                         f"pedais (default: {XLSX_DEFAULT}). Use '' pra desativar.")
    ap.add_argument("--dry-run", action='store_true',
                    help="Mostra o que faria, sem mover nada")
    ap.add_argument("--archive-zip", action='store_true',
                    help="Move o .zip pra subpasta 'processed/' após processar "
                         "(padrão: deixa onde está; reexecutar é seguro pq cada "
                         "arquivo já no destino é pulado).")
    args = ap.parse_args()

    args.exports.mkdir(parents=True, exist_ok=True)
    if not args.dry_run:
        args.pedais.mkdir(parents=True, exist_ok=True)
    handles = load_handles(args.handles)
    sheet_pedais = []
    if args.xlsx and str(args.xlsx) and args.xlsx.exists():
        sheet_pedais = load_pedais_from_xlsx(args.xlsx)
    elif args.xlsx and str(args.xlsx):
        print(f"⚠️  planilha não encontrada: {args.xlsx} — usando só pastas existentes",
              file=sys.stderr)
    print(f"📂 pedais: {args.pedais}")
    print(f"📥 exports: {args.exports}")
    print(f"👥 {len(handles)} handles carregados")
    print(f"📊 {len(sheet_pedais)} pedais lidos da planilha"
          f"{' (' + str(args.xlsx) + ')' if sheet_pedais else ''}")
    if args.dry_run:
        print("🔍 DRY RUN — nada será movido")

    if args.watch:
        print(f"\n👀 vigiando {args.exports} (Ctrl+C pra parar)…")
        seen = set()
        try:
            while True:
                for z in sorted(args.exports.glob('*.zip')):
                    if z in seen:
                        continue
                    seen.add(z)
                    try:
                        process_zip(z, args.pedais, args.exports, handles,
                                    sheet_pedais, args.max_days, args.dry_run,
                                    args.archive_zip)
                    except Exception as e:
                        print(f"  ❌ falha em {z.name}: {e}")
                time.sleep(15)
        except KeyboardInterrupt:
            print("\n👋")
            return

    zips = [Path(z) for z in args.zips] if args.zips else \
           sorted(args.exports.glob('*.zip'))
    if not zips:
        print(f"\n  (nada pra processar em {args.exports})")
        return
    for z in zips:
        try:
            process_zip(z, args.pedais, args.exports, handles,
                        sheet_pedais, args.max_days, args.dry_run,
                        args.archive_zip)
        except Exception as e:
            print(f"  ❌ falha em {z.name}: {e}")

if __name__ == "__main__":
    main()
