#!/usr/bin/env python3
"""
build-photos.py — lê o GPS do EXIF das fotos, gera thumbnails JPEG
web-friendly e escreve public/photos.json para o app exibir as fotos como
círculos no mapa.

Dois modos de origem:

  1. Pasta única (padrão) — varre raw_imgs/, sem associação a pedal:
       python3 build-photos.py

  2. Pastas por pedal — varre recursivamente a saída do coletor
     (coletor_apple_photos.py), onde cada subpasta é um pedal nomeada
     "AAAA-MM-DD - CÓDIGO - Título". Cada foto fica ligada ao seu pedal:
       python3 build-photos.py --photos-root ~/pedais

Com --photos-root, cada foto em photos.json ganha um campo "ride"
{date, code, name}, e o app consegue mostrar as fotos de cada rota.

Por que thumbnails: as fotos vêm em HEIC (Apple), que a maioria dos
navegadores não exibe. Convertendo para JPEG (corrigindo a orientação EXIF)
elas ficam visíveis no popup ao clicar no círculo.

Dependências:
  pip3 install pillow pillow-heif openpyxl
"""

import argparse
import json
import math
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    import pillow_heif
    pillow_heif.register_heif_opener()
    HEIF_OK = True
except ImportError:
    HEIF_OK = False

from PIL import Image, ImageOps, ExifTags

REPO = Path(__file__).resolve().parent.parent
RAW_DEFAULT   = REPO / "raw_imgs"
XLSX_DEFAULT  = REPO / "Censo Hidrográfico - microdados-atualizado.xlsx"
OUT_JSON      = REPO / "public" / "photos.json"
OUT_THUMBS    = REPO / "public" / "photos"
THUMB_MAX_DEF = 1400
THUMB_QUALITY = 82
IMG_EXTS      = {'.heic', '.heif', '.jpg', '.jpeg', '.png'}

# Colunas da aba "Geral" da planilha (1-based) e precedência de código.
COL = dict(data=1, eH=2, PH=3, BP=4, S=5, BT=6, nome=7)
SERIES_PREC = ["PH", "eH", "BT", "BP", "S"]
DATE_RE = re.compile(r'(\d{4}-\d{2}-\d{2})')


def slug(s):
    return re.sub(r'-+', '-', re.sub(r'[^a-z0-9]+', '-', str(s).lower())).strip('-')


def dms_to_decimal(dms, ref):
    """(graus, min, seg) + ref N/S/E/W → grau decimal. None se inválido."""
    try:
        d, m, s = (float(x) for x in dms)
    except (TypeError, ValueError):
        return None
    dec = d + m / 60.0 + s / 3600.0
    if str(ref).strip().upper() in ('S', 'W'):
        dec = -dec
    return dec


def read_exif(img):
    """Devolve {lat, lng, alt, datetime, bearing, fov} — None quando ausente."""
    out = {'lat': None, 'lng': None, 'alt': None, 'datetime': None,
           'bearing': None, 'fov': None}
    try:
        exif = img.getexif()
    except Exception:
        return out
    if not exif:
        return out

    try:
        gps = exif.get_ifd(ExifTags.IFD.GPSInfo)
    except Exception:
        gps = {}
    if gps:
        lat = dms_to_decimal(gps.get(2), gps.get(1))
        lng = dms_to_decimal(gps.get(4), gps.get(3))
        if lat is not None and lng is not None:
            out['lat'] = round(lat, 6)
            out['lng'] = round(lng, 6)
        alt = gps.get(6)
        if alt is not None:
            try:
                out['alt'] = round(float(alt), 1)
            except (TypeError, ValueError):
                pass
        # GPSImgDirection (tag 17) — bússola: para onde a câmera apontava.
        direction = gps.get(17)
        if direction is not None:
            try:
                out['bearing'] = round(float(direction) % 360, 1)
            except (TypeError, ValueError):
                pass

    # Campo de visão — calculado do FocalLengthIn35mmFilm (tag 41989).
    f35 = exif.get(41989)
    if f35:
        try:
            f = float(f35)
            if f > 0:
                # orientações 5–8 = retrato → lado curto do quadro (24 mm)
                portrait = (exif.get(274) or 1) in (5, 6, 7, 8)
                frame = 24.0 if portrait else 36.0
                out['fov'] = round(
                    math.degrees(2 * math.atan(frame / (2 * f))), 1)
        except (TypeError, ValueError):
            pass

    dt_raw = exif.get(36867) or exif.get(306)
    if dt_raw:
        try:
            out['datetime'] = datetime.strptime(
                str(dt_raw).strip(), "%Y:%m:%d %H:%M:%S").isoformat()
        except ValueError:
            pass
    return out


def make_thumbnail(img, dest, max_size):
    """Salva JPEG redimensionado, com orientação EXIF corrigida."""
    img = ImageOps.exif_transpose(img)
    img = img.convert('RGB')
    img.thumbnail((max_size, max_size), Image.LANCZOS)
    img.save(dest, 'JPEG', quality=THUMB_QUALITY, optimize=True)


def load_tour_index(xlsx):
    """Lê a aba Geral → {data_iso: {'code': 'PH 88', 'name': '...'}}."""
    index = {}
    if not xlsx.exists():
        return index
    try:
        import openpyxl
    except ImportError:
        print("⚠️  openpyxl não instalado — pedais sem código/nome", file=sys.stderr)
        return index
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["Geral"]
    for r in range(2, ws.max_row + 1):
        cell = lambda c: ws.cell(row=r, column=c).value
        d = cell(COL["data"])
        if not isinstance(d, datetime):
            continue
        codes = {}
        for name in ("eH", "PH", "BP", "S", "BT"):
            v = cell(COL[name])
            if isinstance(v, int):
                codes[name] = v
        code = next((f"{n} {codes[n]}" for n in SERIES_PREC if n in codes), None)
        nome = cell(COL["nome"])
        index[d.date().isoformat()] = {
            'code': code,
            'name': str(nome).strip() if nome else None,
        }
    return index


def ride_for(folder_name, tour_index):
    """Deriva o pedal a partir do nome da pasta 'AAAA-MM-DD - CÓDIGO - Título'."""
    m = DATE_RE.match(folder_name.strip())
    if not m:
        return None
    date = m.group(1)
    info = tour_index.get(date)
    if info:
        return {'date': date, 'code': info['code'], 'name': info['name']}
    # Sem planilha: tenta extrair código/nome do próprio nome da pasta.
    parts = [p.strip() for p in folder_name.split(' - ')]
    code, name = None, None
    rest = parts[1:]
    if rest and re.match(r'^(PH|eH|BP|S|BT) \d+$', rest[0]):
        code = rest[0]
        rest = rest[1:]
    if rest:
        name = ' - '.join(rest)
    return {'date': date, 'code': code, 'name': name}


def collect_files(args):
    """Devolve lista de (src_path, ride_dict_or_None)."""
    if args.photos_root:
        root = args.photos_root
        if not root.exists():
            sys.exit(f"ERRO: {root} não existe")
        tour_index = load_tour_index(args.xlsx)
        out = []
        for src in sorted(root.rglob('*')):
            if not src.is_file() or src.suffix.lower() not in IMG_EXTS:
                continue
            ride = ride_for(src.parent.name, tour_index) \
                if src.parent != root else None
            out.append((src, ride))
        return out
    # Modo pasta única (raw_imgs/), sem associação a pedal.
    if not args.raw.exists():
        sys.exit(f"ERRO: {args.raw} não existe")
    return [(p, None) for p in sorted(args.raw.iterdir())
            if p.is_file() and p.suffix.lower() in IMG_EXTS]


def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--raw", type=Path, default=RAW_DEFAULT,
                    help=f"pasta única com as fotos (default: {RAW_DEFAULT})")
    ap.add_argument("--photos-root", type=Path, default=None,
                    help="pasta com subpastas por pedal (saída do coletor)")
    ap.add_argument("--xlsx", type=Path, default=XLSX_DEFAULT,
                    help="planilha Censo Hidrográfico (código/nome dos pedais)")
    ap.add_argument("--out-json", type=Path, default=OUT_JSON)
    ap.add_argument("--out-thumbs", type=Path, default=OUT_THUMBS)
    ap.add_argument("--max-size", type=int, default=THUMB_MAX_DEF,
                    help=f"lado maior do thumbnail em px (default: {THUMB_MAX_DEF})")
    args = ap.parse_args()

    args.out_thumbs.mkdir(parents=True, exist_ok=True)
    files = collect_files(args)

    heic = sum(1 for p, _ in files if p.suffix.lower() in ('.heic', '.heif'))
    if heic and not HEIF_OK:
        print(f"⚠️  {heic} arquivos HEIC mas pillow-heif não instalado — "
              f"`pip3 install pillow-heif`", file=sys.stderr)

    src_label = args.photos_root if args.photos_root else args.raw
    print(f"📷 {len(files)} imagens em {src_label}")

    photos, no_gps, failed, used_names = [], 0, 0, set()

    for src, ride in files:
        try:
            img = Image.open(src)
        except Exception as e:
            print(f"  ⚠️  não abriu {src.name}: {e}")
            failed += 1
            continue

        meta = read_exif(img)
        if meta['lat'] is None or meta['lng'] is None:
            no_gps += 1
            continue

        # Nome do thumbnail — único mesmo varrendo várias pastas.
        if args.photos_root:
            base = f"{slug(src.parent.name)[:48]}__{src.stem}"
        else:
            base = src.stem
        thumb_name = base + '.jpg'
        n = 1
        while thumb_name in used_names:
            thumb_name = f"{base}_{n}.jpg"
            n += 1
        used_names.add(thumb_name)

        try:
            make_thumbnail(Image.open(src), args.out_thumbs / thumb_name,
                           args.max_size)
        except Exception as e:
            print(f"  ⚠️  thumbnail falhou {src.name}: {e}")
            failed += 1
            continue

        photos.append({
            'file': f"photos/{thumb_name}",
            'orig': src.name,
            'lat': meta['lat'],
            'lng': meta['lng'],
            'alt': meta['alt'],
            'datetime': meta['datetime'],
            'bearing': meta['bearing'],
            'fov': meta['fov'],
            'ride': ride,
        })
        tag = f"  [{ride['code'] or ride['date']}]" if ride else ""
        print(f"  ✓ {src.name}  →  {meta['lat']}, {meta['lng']}{tag}")

    photos.sort(key=lambda p: p['datetime'] or '')

    payload = {
        'generatedAt': datetime.now().isoformat(timespec='seconds'),
        'count': len(photos),
        'photos': photos,
    }
    args.out_json.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')

    linked = sum(1 for p in photos if p['ride'])
    print(f"\n✅ {len(photos)} fotos georreferenciadas → {args.out_json}")
    print(f"   thumbnails → {args.out_thumbs}/")
    if linked:
        print(f"   ({linked} ligadas a um pedal)")
    if no_gps:
        print(f"   ({no_gps} sem GPS no EXIF, ignoradas)")
    if failed:
        print(f"   ({failed} falharam ao processar)")


if __name__ == "__main__":
    main()
