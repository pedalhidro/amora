#!/usr/bin/env python3
"""
coletor_apple_photos.py — para cada pedal na planilha Censo Hidrográfico,
exporta da biblioteca do Apple Photos as fotos/vídeos feitos na data do
pedal, dentro de uma janela de tempo a partir do horário de início,
organizando-os em pastas por pedal e gravando metadados da ontologia
Pedal Hidrográfico como tags XMP (namespace `ph:`).

Como funciona
  1. Lê a aba "Geral" da planilha: data, horário de início e códigos
     (eH/PH/BP/S/BT) de cada pedal.
  2. Carrega a biblioteca do Photos via osxphotos e casa as fotos por
     janela de tempo. A janela usa o relógio LOCAL da foto, então um
     pedal feito em outro fuso (ex. Oakland) continua casando certo.
  3. Para cada pedal, exporta as fotos via `osxphotos export`: primeiro
     o pedal inteiro de uma vez (rápido); se o osxphotos relatar erro,
     re-exporta foto a foto as que falharam (baixar muitas fotos do
     iCloud de uma vez faz o Photos.app/iCloud falhar — uma a uma, não).
  4. Grava as tags XMP-ph (exiftool) na pasta de cada pedal.

Por que delegar ao `osxphotos export`
  O CLI do osxphotos é o caminho testado para exportar: baixa originais
  do iCloud de forma robusta (`--download-missing` / `--use-photokit`),
  resolve colisões de nome e, com `--update`, pula o que já foi
  exportado. Este script só decide *quais* UUIDs vão para *qual* pasta —
  a planilha continua sendo a fonte da verdade sobre quando cada pedal
  acontece. O casamento por janela é feito aqui (e não pelo
  `--from-date`/`--to-date` do osxphotos) de propósito: ele usa o
  relógio local da foto, o que mantém certo o casamento de pedais em
  outro fuso.

Requisitos (macOS)
  pip3 install osxphotos openpyxl
  brew install exiftool          # necessário p/ gravar os metadados XMP
  Conceda "Acesso Total ao Disco" ao app que roda o script (Terminal/iTerm)
  em  Ajustes do Sistema → Privacidade e Segurança → Acesso Total ao Disco,
  senão a biblioteca do Photos fica ilegível.

Uso
  python coletor_apple_photos.py --list-tours          # só lista os pedais
  python coletor_apple_photos.py --out ~/pedais --dry-run
  python coletor_apple_photos.py --out ~/pedais
  python coletor_apple_photos.py --out ~/pedais --window-hours 7
  python coletor_apple_photos.py --out ~/pedais --download-missing
  python coletor_apple_photos.py --out ~/pedais --no-update
  python coletor_apple_photos.py --out ~/pedais --script export.sh

Notas
  --window-hours 6 (padrão) cobre pedais que passam da meia-noite.
  --download-missing baixa do iCloud os originais ausentes (mais lento).
    Sem essa flag, fotos só-no-iCloud aparecem como "missing" no resumo
    do osxphotos e não são copiadas. Por padrão usa o caminho via
    AppleScript; --use-photokit troca p/ o PhotoKit (mais rápido, mas
    falha com ValueError em algumas fotos HEIC).
  --retry (padrão 3) é o nº de tentativas do osxphotos por foto no passo
    de re-exportação individual. --verbose repassa --verbose ao osxphotos
    e mostra o erro real de cada foto que falha.
  --update (padrão) torna re-execuções incrementais — o osxphotos pula o
    que já foi exportado. Vem acompanhado de --ignore-signature para que
    as tags XMP gravadas por este script sobrevivam às re-execuções (sem
    isso, o osxphotos veria os arquivos como alterados e os sobrescreveria).
    Use --no-update para forçar a re-exportação completa.
  --script ARQUIVO.sh não exporta nada — gera um script bash com uma
    chamada `osxphotos export` por foto p/ você rodar à mão no Terminal.
    Útil p/ depurar (cada linha roda isolada) e p/ tirar o Python do
    meio do processo de exportação.
"""

import argparse
import os
import re
import shlex
import shutil
import subprocess
import sys
import tempfile
from collections import defaultdict
from datetime import datetime, timedelta, time as dtime
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
XLSX_DEFAULT   = REPO / "Censo Hidrográfico - microdados-atualizado.xlsx"
CONFIG_DEFAULT = Path(__file__).resolve().parent / "exiftool_ph.config"
ID_BASE    = "https://pedalhidrografi.co/id/tour/"
COLLECTIVE = "Pedal Hidrográfico"

# Colunas da aba "Geral" (1-based)
COL = dict(data=1, eH=2, PH=3, BP=4, S=5, BT=6, nome=7, horario=19)
# Precedência do código exibido (igual ao app e à ontologia)
SERIES_PREC = ["PH", "eH", "BT", "BP", "S"]

# Extensões de mídia — usadas para contar o que foi exportado e para
# restringir o exiftool aos arquivos de foto/vídeo (e não tocar no
# .osxphotos_export.db que o --update deixa na pasta).
MEDIA_EXT = ("jpg", "jpeg", "heic", "heif", "png", "tiff", "tif",
             "dng", "raw", "cr2", "cr3", "nef", "arw",
             "mov", "mp4", "m4v", "gif")


def sanitize(s):
    """Texto seguro para nome de pasta."""
    s = re.sub(r'[\\/:*?"<>|\n\r\t]+', "-", str(s)).strip()
    s = re.sub(r'\s+', " ", s)
    return s[:90].strip(" .-")


def read_tours(xlsx):
    """Lê a aba Geral → lista de dicts {date, title, ride_code, codes, start}."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["Geral"]
    tours = []
    for r in range(2, ws.max_row + 1):
        cell = lambda c: ws.cell(row=r, column=c).value
        d, nome = cell(COL["data"]), cell(COL["nome"])
        if d is None and nome is None:
            continue
        if not isinstance(d, datetime):
            continue  # sem data não dá para janelar
        codes = {}
        for name in ("eH", "PH", "BP", "S", "BT"):
            v = cell(COL[name])
            if isinstance(v, int):
                codes[name] = v
        ride_code = next(
            (f"{n} {codes[n]}" for n in SERIES_PREC if n in codes), None)
        hor = cell(COL["horario"])
        tours.append(dict(
            date=d.date(),
            title=str(nome).strip() if nome else "",
            ride_code=ride_code,
            codes=codes,
            start=hor if isinstance(hor, dtime) else None,
        ))
    return tours


def folder_name(t):
    parts = [t["date"].isoformat()]
    if t["ride_code"]:
        parts.append(t["ride_code"])
    if t["title"]:
        parts.append(t["title"])
    return sanitize(" - ".join(parts))


def label(t):
    return t["ride_code"] or t["title"] or t["date"].isoformat()


def export_flags(args, retry):
    """Flags comuns do `osxphotos export` — sem o subcomando, a pasta nem
    o seletor de fotos. Reaproveitado tanto para rodar o osxphotos direto
    quanto para gerar o script bash (--script)."""
    # --skip-edited: exporta o original (espelha o comportamento antigo,
    # que copiava photo.path); remova p/ exportar também as edições.
    flags = ["--skip-edited"]
    if args.library:
        flags += ["--library", str(args.library)]
    if args.download_missing:
        # --download-missing baixa do iCloud, por padrão, pelo caminho via
        # AppleScript — é o que funciona de forma confiável aqui.
        # --use-photokit (opcional) troca p/ o framework PhotoKit: é mais
        # rápido, mas falha com ValueError em algumas fotos HEIC.
        flags.append("--download-missing")
        if args.use_photokit:
            flags.append("--use-photokit")
    if retry:
        flags += ["--retry", str(retry)]
    if not args.no_update:
        # --update: re-execuções pulam o que já foi exportado.
        # --ignore-signature: imprescindível aqui — sem ele, o osxphotos
        # veria os arquivos já marcados pelo exiftool como "alterados" e
        # os re-exportaria, apagando as tags XMP-ph.
        flags += ["--update", "--ignore-signature"]
    if args.verbose:
        flags.append("--verbose")
    return flags


def osxphotos_cmd(folder, args, selector, retry=0):
    """Monta a linha de comando do `osxphotos export` para um pedal.

    `selector` são os argumentos que escolhem as fotos, p.ex.
    ["--uuid-from-file", arquivo]  ou  ["--uuid", uuid]. `retry` é o nº de
    tentativas que o osxphotos faz por foto (0 = uma só).
    """
    return (["osxphotos", "export", str(folder), *selector]
            + export_flags(args, retry))


def run_osxphotos(cmd):
    """Roda o osxphotos transmitindo a saída ao vivo e capturando-a.
    Devolve (returncode, texto-da-saída)."""
    proc = subprocess.Popen(cmd, stdout=subprocess.PIPE,
                            stderr=subprocess.STDOUT, text=True, bufsize=1)
    lines = []
    for line in proc.stdout:
        sys.stdout.write(line)
        sys.stdout.flush()
        lines.append(line)
    proc.wait()
    return proc.returncode, "".join(lines)


def export_failed(code, out):
    """True se a exportação teve erro — código != 0 ou erro por foto."""
    return code != 0 or "Error exporting photo" in out


def export_tour(photos, folder, args):
    """Exporta os UUIDs de `photos` para `folder` via `osxphotos export`.

    Estratégia adaptativa: 1º tenta o pedal inteiro numa única chamada
    (rápido); se o osxphotos relatar erro em alguma foto, re-exporta foto
    a foto as que falharam. Baixar muitas fotos do iCloud de uma vez faz
    o Photos.app/iCloud falhar (ValueError); uma a uma funciona. Como o
    --update pula o que já saiu, o 2º passo só baixa o que faltou.
    """
    folder.mkdir(parents=True, exist_ok=True)
    fd, uuid_file = tempfile.mkstemp(suffix=".txt", prefix="phidro-uuids-")
    try:
        with os.fdopen(fd, "w") as f:
            f.write("\n".join(p.uuid for p in photos) + "\n")
        # 1º passo: o pedal inteiro de uma vez (sem --retry — falha rápido
        # p/ o 2º passo assumir).
        code, out = run_osxphotos(
            osxphotos_cmd(folder, args, ["--uuid-from-file", uuid_file]))
    finally:
        os.unlink(uuid_file)

    if not export_failed(code, out):
        return

    # 2º passo: re-exporta foto a foto. Prefere as UUIDs que o osxphotos
    # citou nos erros; se não citou nenhuma, refaz o pedal todo (o
    # --update pula o que já saiu, então é barato).
    bad = list(dict.fromkeys(
        re.findall(r"Error exporting photo \(([0-9A-Fa-f-]+):", out)))
    retry_photos = ([p for p in photos if p.uuid in bad] if bad
                    else list(photos))
    print(f"     ↻ erro na exportação em lote — re-exportando "
          f"{len(retry_photos)} foto(s) uma a uma…")
    still_bad = []
    for p in retry_photos:
        code, out = run_osxphotos(
            osxphotos_cmd(folder, args, ["--uuid", p.uuid], retry=args.retry))
        if export_failed(code, out):
            still_bad.append(p.uuid)
    if still_bad:
        print(f"     ⚠️  {len(still_bad)} foto(s) ainda com erro: "
              f"{', '.join(still_bad)}")


def count_media(folder):
    """Conta os arquivos de mídia presentes na pasta (ignora o .db do --update)."""
    if not folder.exists():
        return 0
    return sum(1 for p in folder.iterdir()
               if p.is_file() and p.suffix.lower().lstrip(".") in MEDIA_EXT)


def exif_cmd(config, folder, t):
    """Monta o comando do exiftool que grava as tags XMP-ph na pasta."""
    iri  = ID_BASE + t["date"].isoformat()
    desc = " — ".join(x for x in (t["ride_code"], t["title"]) if x) \
           or t["date"].isoformat()
    cmd = ["exiftool", "-config", str(config), "-overwrite_original",
           "-q", "-api", "NoDups=1",
           f"-XMP-ph:CapturedDuring={iri}",
           f"-XMP-ph:RideName={t['title']}",
           f"-XMP-ph:RideDate={t['date'].isoformat()}",
           f"-XMP-ph:Collective={COLLECTIVE}",
           f"-XMP-dc:Description={desc}",
           f"-XMP-dc:Subject+={COLLECTIVE}"]
    if t["ride_code"]:
        cmd.append(f"-XMP-ph:RideCode={t['ride_code']}")
        cmd.append(f"-XMP-dc:Subject+={t['ride_code']}")
    for name in ("eH", "PH", "BP", "S", "BT"):
        if name in t["codes"]:
            cmd.append(f"-XMP-ph:RideCodes+={name} {t['codes'][name]}")
    # Restringe às extensões de mídia — assim o exiftool não tropeça no
    # .osxphotos_export.db (SQLite) que o --update deixa na pasta.
    for ext in MEDIA_EXT:
        cmd += ["-ext", ext]
    cmd.append(str(folder))
    return cmd


def write_exif(config, folder, t):
    """Grava as tags XMP-ph (e palavras-chave) em todos os arquivos da pasta."""
    res = subprocess.run(exif_cmd(config, folder, t),
                         capture_output=True, text=True)
    if res.returncode != 0:
        print(f"     ⚠️  exiftool: {res.stderr.strip()}")


def _shline(cmd):
    """Junta um comando para o bash. O literal $DEST sai entre aspas mas
    sem escapar, para o bash expandir a variável."""
    return " ".join('"$DEST"' if a == "$DEST" else shlex.quote(a)
                    for a in cmd)


def write_export_script(plan, args, path):
    """Gera um script bash que exporta os pedais sem passar pelo Python.

    `plan` é uma lista de (t, folder, [photos]). Emite uma chamada
    `osxphotos export` por foto — a unidade comprovadamente confiável — e,
    no fim de cada pedal, a chamada do exiftool. Feito para você rodar à
    mão: qualquer linha pode ser executada isolada para depurar.
    """
    flags = export_flags(args, args.retry)
    lines = [
        "#!/usr/bin/env bash",
        "#",
        f"# Gerado por coletor_apple_photos.py em {datetime.now():%Y-%m-%d %H:%M}.",
        "# Exporta as fotos de cada pedal do Apple Photos — uma chamada",
        "# `osxphotos export` por foto. Rode você mesmo, no Terminal.app:",
        f"#     bash {shlex.quote(path.name)}",
        "# Sem 'set -e': o script segue mesmo se uma foto falhar.",
        "#",
        "set -u",
        "",
    ]
    for t, folder, photos in plan:
        lines.append(f"# === {label(t)} — {len(photos)} foto(s) ===")
        lines.append(f"DEST={shlex.quote(str(folder))}")
        lines.append('mkdir -p "$DEST"')
        for p in photos:
            lines.append(_shline(
                ["osxphotos", "export", "$DEST", "--uuid", p.uuid, *flags]))
        if not args.no_exif:
            ecmd = ["$DEST" if a == str(folder) else a
                    for a in exif_cmd(args.config, folder, t)]
            lines.append(_shline(ecmd))
        lines.append("")
    path.write_text("\n".join(lines) + "\n")
    path.chmod(0o755)


def main():
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx", type=Path, default=XLSX_DEFAULT,
                    help="planilha Censo Hidrográfico")
    ap.add_argument("--out", type=Path,
                    help="pasta base onde criar as pastas por pedal")
    ap.add_argument("--library", type=Path, default=None,
                    help="biblioteca do Photos (padrão: a do sistema)")
    ap.add_argument("--window-hours", type=float, default=6.0,
                    help="horas após o início a incluir (padrão: 6)")
    ap.add_argument("--default-start", default="20:00",
                    help="horário usado quando a planilha não traz um (HH:MM)")
    ap.add_argument("--config", type=Path, default=CONFIG_DEFAULT,
                    help="arquivo de config do exiftool (namespace ph)")
    ap.add_argument("--download-missing", action="store_true",
                    help="baixa do iCloud os originais ausentes (mais "
                         "lento; usa o caminho via AppleScript)")
    ap.add_argument("--use-photokit", action="store_true",
                    help="baixa do iCloud via PhotoKit em vez de AppleScript "
                         "(mais rápido, mas falha com algumas fotos HEIC)")
    ap.add_argument("--retry", type=int, default=3,
                    help="tentativas de re-exportar uma foto que falhou "
                         "(padrão: 3)")
    ap.add_argument("--verbose", action="store_true",
                    help="repassa --verbose ao osxphotos — mostra o erro "
                         "real de cada foto que falha")
    ap.add_argument("--no-update", action="store_true",
                    help="força a re-exportação completa (desliga o "
                         "--update incremental do osxphotos)")
    ap.add_argument("--no-exif", action="store_true",
                    help="não gravar metadados XMP")
    ap.add_argument("--dry-run", action="store_true",
                    help="só mostra o que faria, sem exportar")
    ap.add_argument("--script", type=Path, metavar="ARQUIVO.sh",
                    help="não exporta; gera um script bash (uma chamada "
                         "osxphotos por foto) p/ você rodar à mão")
    ap.add_argument("--list-tours", action="store_true",
                    help="lista os pedais e janelas e sai (não toca no Photos)")
    args = ap.parse_args()

    if not args.xlsx.exists():
        sys.exit(f"ERRO: planilha não encontrada: {args.xlsx}")
    try:
        hh, mm = (int(x) for x in args.default_start.split(":"))
        default_start = dtime(hh, mm)
    except Exception:
        sys.exit(f"ERRO: --default-start inválido: {args.default_start}")

    tours = read_tours(args.xlsx)
    print(f"📋 {len(tours)} pedais com data na planilha")

    if args.list_tours:
        for t in tours:
            st = t["start"] or default_start
            end = (datetime.combine(t["date"], st)
                   + timedelta(hours=args.window_hours))
            flag = "" if t["start"] else "  (horário padrão)"
            print(f"  {t['date']}  {st.strftime('%H:%M')}–{end:%H:%M}"
                  f"  {label(t)}{flag}")
        return

    if not args.out:
        sys.exit("ERRO: informe --out (pasta base) ou use --list-tours")

    try:
        import osxphotos
    except ImportError:
        sys.exit("ERRO: osxphotos não instalado — `pip3 install osxphotos`")
    if not shutil.which("osxphotos"):
        sys.exit("ERRO: CLI `osxphotos` não encontrado no PATH — "
                 "`pip3 install osxphotos`")

    exif_ok = (not args.no_exif) and bool(shutil.which("exiftool"))
    if not args.no_exif and not exif_ok:
        print("⚠️  exiftool não encontrado — metadados XMP não serão gravados "
              "(`brew install exiftool`)")
    if exif_ok and not args.config.exists():
        print(f"⚠️  config não encontrada ({args.config}) — XMP não gravado")
        exif_ok = False

    print("📷 carregando a biblioteca do Photos…")
    db = (osxphotos.PhotosDB(str(args.library)) if args.library
          else osxphotos.PhotosDB())

    # Agrupa as fotos pelo dia local — acelera o casamento por janela.
    bucket = defaultdict(list)
    for p in db.photos():
        if p.date is None:
            continue
        wall = p.date.replace(tzinfo=None)   # relógio local da foto
        bucket[wall.date()].append((wall, p))
    print(f"   {sum(len(v) for v in bucket.values())} itens na biblioteca")

    if not args.script:
        args.out.mkdir(parents=True, exist_ok=True)
    used_names, total, selected, tours_hit = set(), 0, 0, 0
    plan = []   # (t, folder, [photos]) — usado pelo modo --script

    for t in tours:
        st = t["start"] or default_start
        start_dt = datetime.combine(t["date"], st)
        end_dt = start_dt + timedelta(hours=args.window_hours)
        cand = (bucket.get(t["date"], [])
                + bucket.get(t["date"] + timedelta(days=1), []))
        matched = [p for (wall, p) in cand if start_dt <= wall < end_dt]
        if not matched:
            continue
        tours_hit += 1
        selected += len(matched)

        fname = folder_name(t)
        while fname in used_names:
            fname += "_2"
        used_names.add(fname)
        folder = args.out / fname

        if args.script:
            plan.append((t, folder, matched))
            print(f"  ✓  {label(t):14s} {len(matched):3d} fotos → {fname}/")
            continue

        if args.dry_run:
            cmd = osxphotos_cmd(folder, args,
                                ["--uuid-from-file", "<lista-uuids>.txt"])
            print(f"  ✓  {label(t):14s} {len(matched):3d} fotos → {fname}/"
                  f"  (dry-run)")
            print(f"       {' '.join(cmd)}")
            continue

        print(f"  →  {label(t):14s} {len(matched):3d} fotos → {fname}/")
        export_tour(matched, folder, args)
        n = count_media(folder)
        if exif_ok and n:
            write_exif(args.config, folder, t)
        total += n

    if args.script:
        write_export_script(plan, args, args.script)
        print(f"\n✅ script gerado: {args.script}")
        print(f"   {selected} fotos em {tours_hit} pedais. Rode no "
              f"Terminal.app:  bash {args.script}")
        return

    if args.dry_run:
        print(f"\n✅ {selected} fotos selecionadas em {tours_hit} pedais "
              f"(dry-run — nada exportado)")
        return

    print(f"\n✅ {tours_hit} pedais processados → {args.out}")
    print(f"   {selected} fotos casadas na planilha; {total} arquivos de "
          f"mídia presentes nas pastas.")
    print("   (veja os resumos do osxphotos acima p/ exportados x missing.)")
    if not exif_ok and not args.no_exif:
        print("   (metadados XMP não gravados — veja avisos acima)")


if __name__ == "__main__":
    main()
